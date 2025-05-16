import re
import os
import subprocess
import importlib
#import paramiko
import time
import pexpect
import sys
from pexpect import pxssh
import getpass
import pandas
from openpyxl import Workbook, load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import warnings
import paramiko
#import xmltodict
#import pprint
#import json

## Global variables
sysfiles_folder="ccl_labs_sysfiles/"
ccl_labs_folder="/homes/svinukonda/ccl_labs/"
pexpect.run("mkdir -p "+sysfiles_folder)

# Description:
# By default, first argument will be treated as input file
def get_user_input_file_as_dictionary(tool_name, argument_position=1, dest_dir=sysfiles_folder+"/", user_interaction_flag="yes"):
    try:
        if type(argument_position).__name__=='int' and sys.argv[argument_position] != None:
            user_input_file = sys.argv[argument_position]
            user_input = import_module_from_file(user_input_file)
        elif type(argument_position).__name__=='str':
            user_input_file = argument_position
            #flag_file_exist = check_file_exists(user_input_file)
            if user_interaction_flag=="yes":
                ch = USER_INPUT_TEMPLATE1(" Do You Want To Edit The File: {}".format(user_input_file),
                                          {'y': 'Yes. Edit --(default)', 'r': 'reload default and open',
                                           'c': 'continue without open'}, tool_name)
                if ch=="y":
                    os.system("vi {}".format(user_input_file))
                elif ch=="r":
                    cmd = "cp {}/{}/user_input_{}.py {}/".format(ccl_labs_folder, tool_name, tool_name, dest_dir)
                    print(cmd)
                    os.system(cmd)

            user_input = import_module_from_file(user_input_file)
            print("user input file:", argument_position)
    except:
        user_input_file = "user_input_{tool_name}.py".format(tool_name=tool_name)
        ret = check_file_exists(sysfiles_folder + user_input_file)
        if ret == False:
            os.system("mkdir -p "+sysfiles_folder)
            os.system("cp {ccl_labs_folder}/{tool_name}/{user_input_file} {sysfiles_folder}/.".
                      format(tool_name=tool_name, user_input_file=user_input_file, sysfiles_folder=sysfiles_folder, ccl_labs_folder=ccl_labs_folder))
        user_input = get_update_reaload_userInputFile(user_input_file, tool_name=tool_name, dest_dir=dest_dir, user_interaction_flag=user_interaction_flag)
    return user_input


def check_file_exists(path):
    result = subprocess.Popen(['ls', '-lt', path], stdout=subprocess.PIPE,
                              stderr=subprocess.PIPE)
    capstr, err = result.communicate()

    result=True
    mat1=re.search("[rwx\-]+\s+\S+\s+\S+\s+\S+", capstr.decode('UTF-8'))
    if mat1==None:
        result=False

    mat1=re.search("No such file or directory", err.decode('UTF-8'))
    if mat1!=None:
        result=False

    return result

def verify_file_exists(path):
    result = subprocess.Popen(['ls', '-lt', path], stdout=subprocess.PIPE,
                              stderr=subprocess.PIPE)
    capstr, err = result.communicate()

    mat1=re.search("[rw\-]+\s+\S+\s+\S+\s+\S+", capstr.decode('UTF-8'))
    if mat1==None:
        assert False, "\n ** [Error] ** File Not Found : "+path

    mat1=re.search("No such file or directory", err.decode('UTF-8'))
    if mat1!=None:
        assert False, "\n ** [Error] ** File Not Found : "+path

def get_file_names_from_folder_path(path, match_extention_pattern=".*"):
    file_selected =[]
    try:
        file_list = os.listdir(path)
    except NotADirectoryError:
        file_selected.append(path)
        return file_selected

    #print("==>> All files in the directory: ", file_list)
    for file in file_list:
        if re.search(match_extention_pattern, file):
            file_selected.append(file)
    return file_selected

def main_heading_list2(*argv):
    max_len=0
    for v in argv:
        if max_len<len(v):
            max_len=len(v)
    max_len=max_len+2
    main_lst=[]
    for v in argv:
        sub_lst = ["*** <<||", v.ljust(max_len), ">>|| ***"]
        main_lst.append(sub_lst)
    df = pandas.DataFrame(main_lst)
    #df.style.set_properties(**{'text-align': 'right'})
    print(df)

def main_heading_list(*argv):
    print("\n\n")
    max_expected_len=80

    max_len=0
    for v in argv:
        if max_len<len(v):
            max_len=len(v)

    if max_len<max_expected_len:
        max_len=max_expected_len

    prefix_str="#### "
    postfix_str=" ####"
    total_str_len=max_len + len(prefix_str) + len(postfix_str)
    print("     "+"#" * total_str_len)
    print("     "+prefix_str.replace(" ","#") + "#" * max_len + postfix_str.replace(" ","#"))
    print("     "+prefix_str+" "*max_len+postfix_str)
    for v in argv:
        print("     "+prefix_str+v.center(max_len)+postfix_str)
    print("     "+prefix_str + " " * max_len + postfix_str)
    print("     "+prefix_str.replace(" ","#") + "#" * max_len + postfix_str.replace(" ","#"))
    print("     "+"#" * total_str_len)


def print_introduction(tool_name, description, developer, release_date):
    lst1 = []
    lst1.append("".ljust(80,"#"))
    lst1.append("{heading:=^80}".format(heading=" CCL AUTOMATION TOOLS "))
    lst1.append("### {label:^72} ###".format(label="@ 2022 - 2026 Juniper Networks, Inc. All rights reserved."))
    lst1.append("".ljust(80,"#"))
    ""
    lst1.append("### {label:^72} ###".format(label="WelCome To"))
    lst1.append("### {label:^72} ###".format(label=tool_name.upper()))
    lst1.append("### {label:.^72} ###".format(label=""))
    lst1.append("### {label:^72} ###".format(label=""))
    lst1.append("### {label:<15}: {descp:<55} ###".format(label="Description", descp=description))
    lst1.append("### {label:<15}: {descp:<55} ###".format(label="Developer", descp=developer))
    lst1.append("### {label:<15}: {descp:<55} ###".format(label="Release", descp=release_date))
    lst1.append("### {label:^72} ###".format(label=""))
    lst1.append("".ljust(80,"#"))
    print("\n\n")
    for v in lst1:
        print("     "+v)
    print("\n")

def main_heading(str1, print_flag=True, pre_chars=""):
    len1 = len(str1)
    print_str="\n"+pre_chars+" ==>> ***********"+("*"*len1)+"***********"
    print_str=print_str+"\n"+pre_chars+" ==>> *** "+(" "*len1)+"               ***"
    print_str=print_str+"\n"+pre_chars+" ==>> *** <<||   "+str1+"   ||>> ***"
    print_str=print_str+"\n"+pre_chars+" ==>> *** "+(" "*len1)+"               ***"
    #print_str=print_str+"\n\n\n"
    print_str=print_str+"\n"+pre_chars+" ==>> ***********"+("*"*len1)+"***********"
    if print_flag==True:
        print(print_str)

    return print_str

def heading(str1, print_flag=True):
    str1= "\n ==>> *** {}".format(str1)
    if print_flag==True:
        print(str1)
    return str1

def heading_list(*argv):
    print("\n\n")
    for v in argv:
        print(" ==>> *** ", v)

def heading_str(str1, pre_spaces=""):
    return "\n{pre_spaces}==>> *** {str1}".format(pre_spaces=pre_spaces, str1=str1)

"""
def get_update_reaload_userInputFile(user_input_file):

    verify_file_exists(user_input_file)

    cwd = os.getcwd()
    m1 = re.search("([\s\S]+)?\.py", user_input_file)
    if m1==None:
        assert False, "Not found the module name"
    mod = m1.groups()[0]

    spec = importlib.util.spec_from_file_location(mod, cwd+"/"+user_input_file)
    user_input = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(user_input)

    #user_input = importlib.import_module(mod)
    #importlib.reload(user_input)

    return user_input
"""

#user_input_file: user has to provide forward-slash(/)
#shared_dir_name: ccl_labs path
def get_update_reaload_userInputFile(user_input_file, shared_dir_name=None, tool_name=None, ch=None, dest_dir=sysfiles_folder+"/", user_interaction_flag="yes"):
    if shared_dir_name==None:
        mat=re.search("user_input_(.+).py",user_input_file)
        shared_dir_name=ccl_labs_folder+mat.groups()[0]

    #user_input_file: value does have path then
    if re.search("\/user_input", user_input_file):
        user_input_file_with_path=user_input_file
        user_input_file = user_input_file_with_path.split("/")[-1]
    else:
        user_input_file_with_path=sysfiles_folder+"/"+user_input_file


    file_flg = check_file_exists(user_input_file_with_path)
    if file_flg==True:
        if ch==None:
            if user_interaction_flag=="yes":
                ch = USER_INPUT_TEMPLATE1(" INPUT FILE {} already Existed.".format(user_input_file_with_path),
                                 {'y': 'Open existing Input File --(default)', 'r': 'reload default and open', 'c': 'continue without open'}, tool_name)
            else:
                ch = 'c'
    else:
        cmd="cp {}/{} {}".format(shared_dir_name, user_input_file, dest_dir)
        print(cmd)
        os.system(cmd)
        if user_interaction_flag == "yes":
            ch = USER_INPUT_TEMPLATE1(
                "{}: Input File is Copied".format(user_input_file_with_path),
                {'y': 'Open Input File  --(default)', 'c': 'continue without open'}, tool_name)
        else:
            ch = 'c'
    if ch == 'r':
        cmd="cp {}/{} {}".format(shared_dir_name, user_input_file, dest_dir)
        print(cmd)
        os.system(cmd)

        #reload_cmd="cp {}/{} {}".format(shared_dir_name, user_input_file, dest_dir)
        #print("reload cmd:",reload_cmd)
        #os.system(reload_cmd)
        os.system("vi {}".format(user_input_file_with_path))
    elif ch == 'c':
        pass
    elif ch == "y":
        os.system("vi {}".format(user_input_file_with_path))

    cwd = os.getcwd()
    m1 = re.search("([\s\S]+)?\.py", user_input_file_with_path)
    if m1==None:
        assert False, "Not found the module name"
    mod = m1.groups()[0]

    spec = importlib.util.spec_from_file_location(mod, cwd+"/"+user_input_file_with_path)
    user_input = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(user_input)

    #user_input = importlib.import_module(mod)
    #importlib.reload(user_input)

    return user_input


def import_module_from_file(user_input_file):

    if re.search("/",user_input_file):
        file_path=user_input_file
    else:
        cwd = os.getcwd()
        file_path = cwd+"/"+user_input_file
    m1 = re.search("([\s\S]+)?\.py", user_input_file)
    if m1==None:
        assert False, "Not found the module name"
    mod = m1.groups()[0]

    spec = importlib.util.spec_from_file_location(mod, file_path)
    user_input = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(user_input)

    return user_input

def INPUT_FROM_USER(input_msg):
    while True:
        input_val =  input("\n\n\t==> {}: ".format(input_msg))
        #ch = USER_INPUT_TEMPLATE1("Confirm Input", {"y": "Confirm --(default)", "n": "no"})
        if input_msg=="":
            continue

        ch = USER_INPUT_TEMPLATE1("Confirm Input", {"y": "Confirm --(default)", "n": "no"})
        if ch=="y":
            break

    return input_val

def USER_INPUT_TEMPLATE1_old(desc, input_dict, toolname="TOOL"):
    heading(desc)

    selected_char = ""
    while True:
        input_msg = "\n\t==> OPTIONS: "
        for key in sorted(input_dict.keys()):
            input_msg = input_msg+"\n\t\t > {}: {}".format(key, input_dict[key])
        input_msg=input_msg+"\n\n\t==> ENTER YOUR CHOICE: "
        ip=input(input_msg)

        for ch in input_dict.keys():
            selected_char = ch

            if ip.lower().startswith('q'):
                print("\n")
                heading("THAK YOU FOR USING "+toolname)
                heading("BYE.. ")
                quit()
            if ip.lower().startswith(ch):
                break
        break

    return selected_char

#Enter will be taken the default key, if the dictionary value contains the string "(default)"
#it returns only key (not value)
def USER_INPUT_TEMPLATE1(desc, input_dict, toolname="TOOL", comment=""):
    heading(desc)

    selected_char = ""

    while True:
        input_msg = "\n\t==> OPTIONS: "
        flag_default_variable_check = False
        if type(input_dict) is dict:
            input_msg = "\n\t==> OPTIONS: "
            for key in sorted(input_dict.keys()):
                input_msg = input_msg+"\n\t\t > {}: {}".format(key, input_dict[key])
                ret_tbl = re.findall("\(default\)", input_dict[key])
                if len(ret_tbl)>0:
                    flag_default_variable_check = True
                    default_key = key
            input_msg = input_msg + "\n\t\t > {}: {}".format("q", "quit")
            input_msg=input_msg+"\n\n\t==> ENTER YOUR CHOICE: "

            ip = input(input_msg)

            if ip=="" and flag_default_variable_check==True:
                selected_char=default_key
                break

            if ip=="":
                continue

            if ip.lower().startswith('q'):
                print("\n")
                main_heading_list("THAK YOU FOR USING " + toolname, "", comment)
                quit()

            if ip[0] in input_dict.keys():
                selected_char=ip[0]
                break


        elif type(input_dict) is list:
            for key in range(len(input_dict)):
                input_msg = input_msg + "\n\t\t > {}: {}".format(key+1, input_dict[key])
                ret_tbl = re.findall("\(default\)", input_dict[key])
                if len(ret_tbl)>0:
                    flag_default_variable_check = True
                    default_key = key
                    default_value = input_dict[key]

            input_msg = input_msg + "\n\t\t > q: Quit"
            input_msg = input_msg + "\n\n\t==> ENTER YOUR CHOICE: "
            ip = input(input_msg)

            if ip.lower().startswith('q'):
                print("\n")
                main_heading_list("THAK YOU FOR USING " + toolname, "", comment)
                quit()
            try:
                if ip=="" and flag_default_variable_check==True:
                    selected_char=int(default_key)+1
                    break

                if int(ip) in range(len(input_dict)+1):
                    selected_char=int(ip)
                    #print("##SUCH selected char", selected_char)
                    break

            except:
                continue

    return selected_char

##INPROGRESS
def USER_INPUT_SEARCH_TEMPLATE1(desc, input_dict, toolname="TOOL", comment=""):
    heading(desc)

    selected_char = ""

    while True:
        input_msg = "\n\t==> OPTIONS: "
        flag_default_variable_check = False
        if type(input_dict) is dict:
            input_msg = "\n\t==> OPTIONS: "
            for key in sorted(input_dict.keys()):
                input_msg = input_msg+"\n\t\t > {}: {}".format(key, input_dict[key])
                ret_tbl = re.findall("\(default\)", input_dict[key])
                if len(ret_tbl)>0:
                    flag_default_variable_check = True
                    default_key = key
            input_msg = input_msg + "\n\t\t > {}: {}".format("q", "quit")
            input_msg=input_msg+"\n\n\t==> ENTER YOUR CHOICE: "

            ip = input(input_msg)

            if ip=="" and flag_default_variable_check==True:
                selected_char=default_key
                break

            if ip=="":
                continue

            if ip.lower().startswith('q'):
                print("\n")
                main_heading_list("THAK YOU FOR USING " + toolname, "", comment)
                quit()

            if ip[0] in input_dict.keys():
                selected_char=ip[0]
                break


        elif type(input_dict) is list:
            for key in range(len(input_dict)):
                input_msg = input_msg + "\n\t\t > {}: {}".format(key+1, input_dict[key])
                ret_tbl = re.findall("\(default\)", input_dict[key])
                if len(ret_tbl)>0:
                    flag_default_variable_check = True
                    default_key = key
                    default_value = input_dict[key]

            input_msg = input_msg + "\n\t\t > q: Quit"
            input_msg = input_msg + "\n\n\t==> ENTER YOUR CHOICE: "
            ip = input(input_msg)

            if ip.lower().startswith('q'):
                print("\n")
                main_heading_list("THAK YOU FOR USING " + toolname, "", comment)
                quit()
            try:
                if len(ip)>1:
                    pass

                if ip=="" and flag_default_variable_check==True:
                    selected_char=int(default_key)+1
                    break

                if int(ip) in range(len(input_dict)+1):
                    selected_char=int(ip)
                    #print("##SUCH selected char", selected_char)
                    break

            except:
                continue

    return selected_char


#Return two values
# one is character, another one is list value
# not tested for dictionary
def USER_INPUT_TEMPLATE2(desc, input_dict, toolname="TOOL", comment="", number_of_colomns=1):
    heading(desc)

    selected_char = ""
    selected_val = ""

    while True:
        input_msg = "\n\t==> OPTIONS: "
        if type(input_dict) is dict:
            input_msg = "\n\t==> OPTIONS: "
            itr=0
            for key in sorted(input_dict.keys()):
                itr=itr+1
                if itr%number_of_colomns==0:
                    input_msg = input_msg+"\n\t\t > {}: {}".format(key, input_dict[key])
                    ret_tbl = re.findall("\(default\)", input_dict[key])
                    if len(ret_tbl)>0:
                        flag_default_variable_check = True
                        default_key = key
                else:
                    input_msg = input_msg+"\t\t > {}: {}".format(key, input_dict[key])

            input_msg = input_msg + "\n\t\t > q: quit"
            input_msg=input_msg+"\n\n\t==> ENTER YOUR CHOICE: "
            ip=input(input_msg)

            if ip.lower().startswith('q'):
                print("\n")
                main_heading_list("THAK YOU FOR USING " + toolname, "", comment)
                quit()

            if ip=="" and flag_default_variable_check==True:
                selected_char=int(default_key)+1
                break

            if ip[0] in input_dict.keys():
                selected_char=ip[0]
                selected_val=input_dict[ip[0]]
                break

        elif type(input_dict) is list:
            itr=0
            for key in range(len(input_dict)):
                itr=itr+1
                if itr % number_of_colomns == 0:
                    input_msg = input_msg + "\n\t\t > {}: {}".format(key+1, input_dict[key])
                else:
                    input_msg = input_msg + "\t\t > {}: {}".format(key+1, input_dict[key])
                ret_tbl = re.findall("\(default\)", input_dict[key])
                if len(ret_tbl) > 0:
                    flag_default_variable_check = True
                    default_key = key
                    default_value = input_dict[key]

            input_msg = input_msg + "\n\t\t > q: Quit"
            input_msg = input_msg + "\n\n\t==> ENTER YOUR CHOICE: "
            ip = input(input_msg)

            if ip.lower().startswith('q'):
                print("\n")
                main_heading_list("THAK YOU FOR USING " + toolname, "", comment)
                quit()
            try:

                if ip == "" and flag_default_variable_check == True:
                    selected_char = int(default_key) + 1
                    selected_val = default_value
                    break

                if int(ip) in range(len(input_dict)+1):
                    selected_char=int(ip)
                    selected_val=input_dict[int(ip)-1]
                    break
            except:
                continue
    #print("##SUCH selected_char, selected_val", selected_char, selected_val)
    return selected_char, selected_val

def CHOOSE_FROM_OPTIONS(desc, input_dict, toolname="TOOL", comment="", number_of_colomns=1):
    heading(desc)

    selected_char = ""
    selected_val = ""

    input_dict_org = input_dict.copy()
    desc_org = desc
    while True:
        input_msg = "\n\t==> OPTIONS: "
        if type(input_dict) is dict:
            input_msg = "\n\t==> OPTIONS: "
            itr=0
            for key in sorted(input_dict.keys()):
                itr=itr+1
                if itr%number_of_colomns==0:
                    input_msg = input_msg+"\n\t\t > {}: {}".format(key, input_dict[key])
                    ret_tbl = re.findall("\(default\)", input_dict[key])
                    if len(ret_tbl)>0:
                        flag_default_variable_check = True
                        default_key = key
                else:
                    input_msg = input_msg+"\t\t > {}: {}".format(key, input_dict[key])

            input_msg = input_msg + "\n\t\t > q: quit"
            input_msg=input_msg+"\n\n\t==> ENTER YOUR CHOICE: "
            ip=input(input_msg)

            if ip.lower().startswith('q'):
                print("\n")
                main_heading_list("THAK YOU FOR USING " + toolname, "", comment)
                quit()

            if ip=="" and flag_default_variable_check==True:
                selected_char=int(default_key)+1
                break

            if ip[0] in input_dict.keys():
                selected_char=ip[0]
                selected_val=input_dict[ip[0]]
                break

        elif type(input_dict) is list:
            itr=0
            for key in range(len(input_dict)):
                itr=itr+1
                if itr % number_of_colomns == 0:
                    input_msg = input_msg + "\n\t\t > {}: {}".format(key+1, input_dict[key])
                else:
                    input_msg = input_msg + "\t\t > {}: {}".format(key+1, input_dict[key])
                ret_tbl = re.findall("\(default\)", input_dict[key])
                if len(ret_tbl) > 0:
                    flag_default_variable_check = True
                    default_key = key
                    default_value = input_dict[key]

            input_msg = input_msg + "\n\t\t > q: Quit"
            input_msg = input_msg + "\n\n\t==> ENTER YOUR CHOICE: "
            ip = input(input_msg)


            if ip.lower().startswith('q'):
                print("\n")
                main_heading_list("THAK YOU FOR USING " + toolname, "", comment)
                quit()


            ######------------------------{{
            reset_toollst_str = "*** This is Sorted Tool List.. Get All Tool List"
            try:
                if input_dict[int(ip)-1]==reset_toollst_str:
                    desc= desc_org
                    input_dict = input_dict_org.copy()
                    continue
            except Exception as err:
                print(err)
                pass

            try:
                if type(int(ip)).__name__ == 'int':
                    pass
            except:
                new_lst = [tool for tool in input_dict if ip in tool]
                if len(new_lst)>0:
                    input_dict = new_lst.copy()
                    if reset_toollst_str not in input_dict:
                        input_dict.append(reset_toollst_str)
                    desc = desc + "--(**SORTED DATA**)"
                    continue
            ######------------------------}}

            try:
                if ip == "" and flag_default_variable_check == True:
                    selected_char = int(default_key) + 1
                    selected_val = default_value
                    break

                if int(ip) in range(len(input_dict)+1):
                    selected_char=int(ip)
                    selected_val=input_dict[int(ip)-1]
                    break
            except:
                continue
    #print("##SUCH selected_char, selected_val", selected_char, selected_val)
    return selected_val


def USER_INPUT_TEMPLATE3(desc, input_dict, toolname="TOOL", comment="", number_of_colomns=1):
    heading(desc)

    selected_char = ""
    selected_val = ""
    print("##SUCH kdkdjf")
    while True:
        input_msg = "\n\t==> OPTIONS: "
        if type(input_dict) is dict:
            input_msg = "\n\t==> OPTIONS: "
            itr=0
            for key in sorted(input_dict.keys()):
                itr=itr+1
                if itr%number_of_colomns==0:
                    input_msg = input_msg+"\n\t\t > {}: {}".format(key, input_dict[key])
                    ret_tbl = re.findall("\(default\)", input_dict[key])
                    if len(ret_tbl)>0:
                        flag_default_variable_check = True
                        default_key = key
                else:
                    input_msg = input_msg+"\t\t > {}: {}".format(key, input_dict[key])

            input_msg = input_msg + "\n\t\t > q: quit"
            input_msg=input_msg+"\n\n\t==> ENTER YOUR CHOICE: "
            ip=input(input_msg)

            if ip.lower().startswith('q'):
                print("\n")
                main_heading_list("THAK YOU FOR USING " + toolname, "", comment)
                quit()

            if ip=="" and flag_default_variable_check==True:
                selected_char=int(default_key)+1
                break

            if ip[0] in input_dict.keys():
                selected_char=ip[0]
                selected_val=input_dict[ip[0]]
                break



        elif type(input_dict) is list:
            itr=0
            for key in range(len(input_dict)):
                itr=itr+1
                if itr % number_of_colomns == 0:
                    input_msg = input_msg + "\n\t\t > {}: {}".format(key+1, input_dict[key])
                else:
                    input_msg = input_msg + "\t\t > {}: {}".format(key+1, input_dict[key])
                ret_tbl = re.findall("\(default\)", input_dict[key])
                if len(ret_tbl) > 0:
                    flag_default_variable_check = True
                    default_key = key
                    default_value = input_dict[key]

            input_msg = input_msg + "\n\t\t > q: Quit"
            input_msg = input_msg + "\n\n\t==> ENTER YOUR CHOICE: "
            ip = input(input_msg)


            if re.search("\w+", ip):
                print("##SUCH in 1")
                lst = re.findall("\w+", ip)
                selected_val=ip
                for itm in lst:
                    print("##SUCH in 2")
                    num=re.search("(\w+)", itm)
                    print("##SUCH num", num)
                    selected_val = re.sub(itm, input_dict[int(num.groups()[0])-1], selected_val)
                break

            if ip.lower().startswith('q'):
                print("\n")
                main_heading_list("THAK YOU FOR USING " + toolname, "", comment)
                quit()
            try:
                print("##SUCH kdkdjf")

                if ip == "" and flag_default_variable_check == True:
                    selected_char = int(default_key) + 1
                    selected_val = default_value
                    break

                if int(ip) in range(len(input_dict)+1):
                    selected_char=int(ip)
                    selected_val=input_dict[int(ip)-1]
                    break

                print("##SUCH kdkdjf")
                if re.search("<\w+>", ip):
                    print("##SUCH in list")
                    lst = re.findall("<\w+>", ip)
                    selected_val=ip
                    for itm in lst:
                        num=re.search("(\w+)", itm)
                        print("##SUCH num", num)
                        selected_val = re.sub(itm, input_dict[num.groups()[0]], selected_val)
                    break

            except:
                continue
    #print("##SUCH selected_char, selected_val", selected_char, selected_val)
    return selected_val


def get_user_input_info(shared_folder_user_input_file_path, tool_name, arg_num=1):

    user_input_file = shared_folder_user_input_file_path.split("/")[-1]
    shared_folder_path = shared_folder_user_input_file_path.split(user_input_file)[0]

    try:
        if sys.argv[1] != None:
            user_input_file = sys.argv[arg_num]
    except:
        ret = check_file_exists(user_input_file)
        if ret == False:
            os.system("cp {} .".format(shared_folder_user_input_file_path))

        uip = USER_INPUT_TEMPLATE1(
            "NOTE: INPUT FILE: Please provide the Info In File: {}".format(user_input_file),
            {
                'y': 'yes',
                'c': 'continue without opening',
                'r': 'reload the default file and open',
                'q': 'exit'
            },
            "GET INVEROTY INFO TOOL"
        )
        if uip == 'y':
            os.system("nano {}".format(user_input_file))
        elif uip == 'c':
            pass
        elif uip == 'r':
            os.system("cp {} .".format(shared_folder_user_input_file_path))
            os.system("nano {}".format(user_input_file))


    #user_input_file = "get-inventory-info-input.py"
    user_input=get_update_reaload_userInputFile(user_input_file)

    return user_input


def get_folderpath_from_publichtml(public_folder_path):
    result = subprocess.Popen(['readlink', '-f', public_folder_path], stdout=subprocess.PIPE,
                              stderr=subprocess.PIPE)
    capstr1, err = result.communicate()

    result = subprocess.Popen(['ls', '-lt', capstr1.strip()], stdout=subprocess.PIPE,
                              stderr=subprocess.PIPE)
    capstr2, err = result.communicate()
    if len(str(err)) > 3:
        warnings.warn("\n\n ==> " + str(err))
        return -1

    toby_folder_path = capstr1.decode().strip().strip("\n")
    return toby_folder_path

def get_folderpath_from_linkname(link_name, dir_path="./"):
    result = subprocess.Popen(['readlink', '-f', dir_path+"/"+link_name], stdout=subprocess.PIPE,
                              stderr=subprocess.PIPE)
    capstr1, err = result.communicate()

    result = subprocess.Popen(['ls', '-lt', capstr1.strip()], stdout=subprocess.PIPE,
                              stderr=subprocess.PIPE)
    capstr2, err = result.communicate()
    if len(str(err)) > 3:
        warnings.warn("\n\n ==> " + str(err))
        cmd_exe= "readlink -f {}/{} > {}".format(dir_path, link_name, sysfiles_folder+"readlink_output")
        os.system(cmd_exe)
        with open(sysfiles_folder+"readlink_output", "r") as fd:
            capstr1 = fd.read()
            print(capstr1)
        if len(capstr1) <= 3:
            return -1
        toby_folder_path = capstr1.strip().strip("\n")
    else:

        toby_folder_path = capstr1.decode().strip().strip("\n")
    return toby_folder_path
def create_public_link(source_dir_from_cwd, folder_in_public_html, symbolic_link_name):
    start_cwd = os.getcwd()
    os.chdir(source_dir_from_cwd)
    source_dir_from_cwd = os.getcwd()
    cwd = source_dir_from_cwd
    user_name=os.environ['USER']
    lst1= symbolic_link_name.split("/")
    while ("" in lst1):
        lst1.remove("")

    #commenting following lines (it was doing chmod on highlevel folder and taking more time)
    #cmd="chmod -R 755 /{}/{}/{} 2>/dev/null".format(cwd.split("/")[1],cwd.split("/")[2],cwd.split("/")[3])
    #print(cmd)
    #os.system(cmd)

    os.system("chmod -R 755 ~/public_html")

    os.system("mkdir -p ~/public_html/"+ folder_in_public_html)
    cmd="chmod -R 755 ~/public_html/" + folder_in_public_html
    os.system(cmd)
    cmd="cd ~/public_html/{}; ln -sfn {}/ {}".format(folder_in_public_html, cwd, lst1[-1])
    os.system(cmd)
    cmd="chmod -R 777 {}".format(cwd)
    #print(cmd)
    #os.system(cmd)
    link_name="https://ttsv-web01.juniper.net/~{}/{}/{}".format(user_name, folder_in_public_html,lst1[-1])
    heading("created Link: "+link_name)
    heading("Pleae execute command (incase of permission deny): chmod -R 777 "+start_cwd)
    return link_name

def create_default_pexpect_spawn(cmd="bash"):

    child = pexpect.spawn('/bin/bash', encoding='utf-8')
    #child.logfile_read = sys.stdout
    child.logfile = sys.stdout

    return child

def pexpect_working_for_reference():
    cwd=""
    child = pexpect.spawn("bash", encoding='utf-8')
    child.logfile_read = sys.stdout
    child.logfile = sys.stdout
    print("##XX##", child.before)

    child.sendline("cd " + cwd)
    child.expect_exact("$ ")

    child.sendline("ls -lt")
    child.expect_exact("$ ")
    child.interact()

    print("##XX##", child.before)

def read_file_to_list(file_name):
	with open(file_name,"r") as f:
	    lines=f.readlines()
	    return lines


def ensure_remote_dir(sftp, remote_dir):
    print("remote_dir", remote_dir)
    """Recursively create directories if they do not exist on the remote machine."""
    dirs = remote_dir.replace("\\", "/").split("/")  # Handle Windows & Linux paths
    path = ""

    for directory in dirs:
        if directory:  # Skip empty parts
            path += "/" + directory + "/"
            try:
                print("path", path)
                sftp.chdir(path)  # Check if directory exists
            except IOError:
                print(f"Creating directory: {path}")
                sftp.mkdir(path)  # Create missing directory
                sftp.chdir(path)  # Move into newly created directory

def scp_directory(hostname, username, password, local_dir, remote_dir, port=22):
    #try:
    # Create SSH client
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname, port=port, username=username, password=password)

    # Start SFTP session
    sftp = ssh.open_sftp()

    # Create nested directories recursively
    ensure_remote_dir(sftp, remote_dir)

    # Function to recursively upload a directory
    def upload_dir(local_path, remote_path):
        try:
            sftp.mkdir(remote_path)  # Create directory on remote machine
            print(f"Created remote directory: {remote_path}")
        except:
            pass  # Ignore if directory exists

        print("local_path", local_path)
        for item in os.listdir(local_path):
            local_item = os.path.join(local_path, item)
            remote_item = os.path.join(remote_path, item).replace("\\", "/")  # Ensure proper path format

            if os.path.isdir(local_item):  # If it's a directory, recurse
                upload_dir(local_item, remote_item)
            else:  # If it's a file, upload it
                sftp.put(local_item, remote_item)
                print(f"Uploaded: {local_item} -> {remote_item}")

    # Start copying
    upload_dir(local_dir, remote_dir)

    # Close connections
    sftp.close()
    ssh.close()

    print("Directory transfer successful!")

    #except Exception as e:
    #    print(f"File transfer failed: {e}")


def scp_file(hostname, username, password, local_file, remote_file, port=22):
    try:
        # Create SSH client
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(hostname, port=port, username=username, password=password)

        # Start SFTP session
        sftp = ssh.open_sftp()

        # Upload file
        sftp.put(local_file, remote_file)
        print(f"File {local_file} successfully copied to {hostname}:{remote_file}")

        # Close connections
        sftp.close()
        ssh.close()

    except Exception as e:
        print(f"File transfer failed: {e}")

def ssh_windows(hostname, username, password, timeout=30):
    try:
        # Create SSH client
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        # Connect to Windows SSH server
        client.connect(hostname, username=username, password=password, timeout=timeout)

        ## Execute commands
        #commands = ["systeminfo", "date /T", "time /T"]
        #for cmd in commands:
        #    stdin, stdout, stderr = client.exec_command(cmd)
        #    output = stdout.read().decode().strip()
        #    print(f"Output of '{cmd}':\n{output}")
        #
        ## Close connection
        #client.close()

    except Exception as e:
        print(f"SSH connection error: {e}")

    return client

def ssh_login(hostname, username, password, timeout=7200, prompt=None):
    try:
        heading("ssh login: "+hostname)
        s = pxssh.pxssh(timeout=timeout)
        s.login(hostname, username, password)
        s.sendline('uptime')  # run a command
        s.prompt()  # match the prompt
        print(s.before.decode('UTF-8'))  # print everything before the prompt.
        s.sendline('date')
        if prompt!=None:
            s.PROMPT = prompt
        s.prompt()
        print(s.before.decode('UTF-8'))
        #s.sendline('df')
        #s.prompt()
        #print(s.before.decode('UTF-8'))
    except pxssh.ExceptionPxssh as e:
        print("pxssh failed on login.")
        print(e)
    return s

def remove_duplicates_in_list(tag_list1):
    tag_list2 = []
    for i in tag_list1:
        if i not in tag_list2:
            tag_list2.append(i)
    return tag_list2


def create_workbook(wb_name, active_sheet_rename=None):
    wb = Workbook()
    ws = wb.active
    if active_sheet_rename !=None:
        ws.title="summary"
    wb.save(wb_name)


def create_workbook_sheet(wb_name, sheet_name):
    wb = load_workbook(wb_name)
    if sheet_name in wb.sheetnames:
        assert False, "Sheet Name Alredy Existed"
    wb.create_sheet(sheet_name)
    wb.save(wb_name)

def create_workbook_sheet_if_not_exist(wb_name, sheet_name):
    wb = load_workbook(wb_name)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    wb.save(wb_name)

def add_value_to_workbook_sheet(wb_name, sheet_name, row_num, col_num, value, result=True):
    wb = load_workbook(wb_name)
    ws = wb[sheet_name]

    d = ws.cell(row=int(row_num), column=int(col_num), value=value)
    ft = Font(color="FF0000")

    if result!=True:
        d.font = ft

    wb.save(wb_name)

def update_values_to_workbook_sheet(wb_name, sheet_name, xls_data):
    wb = load_workbook(wb_name)
    ws = wb[sheet_name]
    ft = Font(color="FF0000")

    for lst in xls_data:
        d = ws.cell(row=int(lst[0]), column=int(lst[1]), value=lst[2])
        if lst[3]!=True:
            d.font = ft

    wb.save(wb_name)



def get_input_from_user(input_desc, default_val=None, convert_input_type_to="str", wait_for_confirmation="no"):
    while True:

        if default_val==None:
            ip_str = input("\n\n==>> ***  {}: ".format(input_desc))

        if default_val != None:
            while True:
                ip_tmp=input("\n\n==>> ***  {} [: default:: {}]: ".format(input_desc, default_val))
                if ip_tmp=="":
                    ip_str=default_val
                else:
                    ip_str=ip_tmp

                break

            if ip_tmp == "":
                return ip_str

        try:
            if convert_input_type_to=="int":
                print("int input")
                ip_str=int(ip_str)
            elif convert_input_type_to == "float":
                    print("float input", ip_str)
                    ip_str = float(ip_str)
            elif convert_input_type_to=="str":
                pass
            else:
                assert False, "Valid input types are: int and str"
        except:
            print("\t==>> ***ERROR*** Not able to convert input to {}".format(convert_input_type_to))
            continue

        if wait_for_confirmation=="yes":
            if default_val==None:
                ch=USER_INPUT_TEMPLATE1("Confirm?", {"y": "Yes  --(default)", "c": "cancel"})
                if ch=="y":
                    break
            else:
                ch=USER_INPUT_TEMPLATE1("Confirm?", {"y": "Yes  --(default)", "c": "cancel"})
                if ch=="y":
                    break
        else:
            break

    return ip_str